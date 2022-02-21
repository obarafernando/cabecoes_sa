import boto3
import logging
import io
from openpyxl import load_workbook
import datetime
import json

FORMAT = '%(asctime)-15s - %(message)s'
logging.basicConfig(format=FORMAT)
log = logging.getLogger()
log.setLevel(logging.INFO)

def list_s3_keys(bucket_name, prefix):
    S3 = boto3.client('s3')
    log.info(f'Listing files in s3://{bucket_name}/{prefix}')
    bucket_paginator = S3.get_paginator('list_objects_v2').paginate(Bucket=bucket_name, Prefix=prefix)
    for page in bucket_paginator:
        for f in page.get('Contents', []):
            yield f['Key']

def fetch_last_run_datetime(bucket_name, prefix, checksum_separator=None):
    valid_datetimes = []
    for key in list_s3_keys(bucket_name, prefix):
        try:
            datetime_raw = key.split('/')[-1].split('.')[0]
            if checksum_separator:
                datetime_raw = datetime_raw.split(checksum_separator)[0]
            valid_datetimes.append(datetime_raw)
        except (ValueError, IndexError) as e:
            log.warning(f'Got error from parsing past runs: {e}')
    if not valid_datetimes:
        return None
    return max(valid_datetimes)

def s3_binary_to_workbook(file,table_name):
    data = io.BytesIO(file['Body'].read())
    table = load_workbook(data).active
    table = convert_none_null(convert_dt_columns_values(table,table_name))
    return table

def add_info_columns(tables):
    for table, last_dt, table_name in tables:
        try:
            load_time = datetime.datetime.now()
            load_time = load_time.strftime("%Y-%m-%d %H:%M:%S")
            last_row = len(table['A'])
            table.insert_cols(idx=1, amount=2)
            table['A1'] = 'FILE_DT'
            table['B1'] = 'LOAD_TIME'
            for x in range(2, last_row):
                table[f'A{x}'] = last_dt
                table[f'B{x}'] = load_time
            yield table, table_name
        except:
            pass

def convert_dt_columns_values(table,table_name):
    max_row = len(table['A'])
    max_col = table.max_column
    if table_name == 'fato_leitura':
        for rows in range(2,max_row):
            for columns in range(max_col-2 ,max_col+1):
                cell = table.cell(row=rows, column=columns)
                cell.value = str(cell.value)
    return table

def convert_none_null(table):
    max_row = len(table['A'])
    max_col = table.max_column
    for rows in range(2,max_row):
        for columns in range(1,max_col+1):
            cell = table.cell(row=rows, column=columns)
            if cell.value == 'NULL' or cell.value == None:
                cell.value = ''
            if cell.value == '0.000000000000000000':
                cell.value = 0
    return table

def return_table_schema(table_name, types):
    with open('cabecoes_sa\schema.json') as json_file:
        schema = json.load(json_file)
        schema_string = ''
        if table_name in schema:
            for column in schema[table_name]:
                if types == 'types':
                    schema_string += column['name']+' '+column['type']+','
                else:
                    schema_string += column['name']+','
            schema_string = schema_string[:-1]
        return schema_string